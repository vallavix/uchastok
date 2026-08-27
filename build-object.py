#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Собирает файл объекта для приложения «Участок».

В файл входит:
  - проект: типы квартир со спецификациями ЭОМ (из рабочих .xlsx);
  - раскладка: этажность, квартир на этаже и кто на каком месте по стояку;
  - данные обхода (необязательно): что принято, замечания и объёмы по бригадам
    из резервной копии «Электро-приёмки». Фотографии не берутся — они не нужны
    начальнику и раздувают файл в сотни раз.

Запуск:
    python3 build-object.py

Пути правятся ниже, в блоке НАСТРОЙКИ.
"""

import json
import os
import re
import sys

# ============ НАСТРОЙКИ ============

OBJECT_NAME = 'ЖК Химки-тайм'
PROJECTS_DIR = os.path.expanduser('~/Desktop/проекты')
# резервная копия из «Электро-приёмки»; None — собрать файл без данных обхода
WALK_BACKUP = os.path.expanduser('~/Desktop/Шахматка_копия_2026-08-27.json')
OUT = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'ЖК_Химки_объект.json')

# какие листы берём в каждый корпус (шифры ЭОМ)
BUILDINGS = [
    {
        'id': 'b1', 'name': 'Корпус 1',
        'xlsx': 'к1/Рабочая_Химки-КВАРТИРЫ к.1 (1).xlsx',
        'types': ['ФП-1.1', 'ФП-1.2', 'ФП-1.2.1', 'ФП-1.3', 'ФП-1.С', 'ФП-1.С.1'],
        # этажи 2-20, 10 квартир на этаже, нумерация с №1 на втором этаже
        'preset': {
            'from': 2, 'to': 20, 'per': 10, 'first': 1,
            'stack': ['ФП-1.С', 'ФП-1.С', 'ФП-1.2', 'ФП-1.1', 'ФП-1.2.1',
                      'ФП-1.2.1', 'ФП-1.1', 'ФП-1.2', 'ФП-1.С', 'ФП-1.С'],
            # наверху меняются только средние
            'ex': [{'from': 19, 'to': 20, 'per': 10,
                    'stack': ['ФП-1.С', 'ФП-1.С', 'ФП-1.3', 'ФП-1.С.1', 'ФП-1.2.1',
                              'ФП-1.2.1', 'ФП-1.С.1', 'ФП-1.3', 'ФП-1.С', 'ФП-1.С']}],
        },
    },
    {
        'id': 'b2', 'name': 'Корпус 2',
        'xlsx': 'к2/Рабочая_Химки-КВАРТИРЫ к.2.xlsx',
        # 1.С и 1.1 лежат в файле корпуса 2 — типовые проекты общие для обоих корпусов
        'types': ['ФП-1.С', 'ФП-1.1', 'ФП-2.С', 'ФП-2.2', 'ФП-2.1', 'ФП-2.3'],
        'preset': {
            'from': 2, 'to': 20, 'per': 12, 'first': 1,
            'stack': ['ФП-1.С', 'ФП-1.С', 'ФП-2.С', 'ФП-2.2', 'ФП-1.1', 'ФП-2.1',
                      'ФП-2.1', 'ФП-1.1', 'ФП-2.2', 'ФП-2.С', 'ФП-1.С', 'ФП-1.С'],
            # угловые 2.С становятся 2.3, а 2.2 пропадает — отсюда 10 вместо 12
            'ex': [{'from': 18, 'to': 20, 'per': 10,
                    'stack': ['ФП-1.С', 'ФП-1.С', 'ФП-2.3', 'ФП-1.1', 'ФП-2.1',
                              'ФП-2.1', 'ФП-1.1', 'ФП-2.3', 'ФП-1.С', 'ФП-1.С']}],
        },
    },
]

# Приход материалов на объект, по накладным. Ключ — начало названия позиции
# в спецификации, значение — сколько пришло. Заполняется по бумажному учёту;
# дальше начальник правит прямо в приложении, в колонке «Пришло».
RECEIVED = {
    'b1': {
        'Рамка 1-я': 724,
        'Рамка 2-я': 360,
        'Рамка 3-я': 148,
        'Рамка 4-я': 140,
        'Выключатель одноклавишный': 800,
        'Выключатель двухклавишный': 267,
        'Светильник класса защиты': 196,
    },
}

# позиции подсчёта из обхода → группы приложения
GROUPS = [
    ('socket', r'розетк'), ('swtch', r'выключател'), ('light', r'светильник|люстр'),
    ('lamp', r'бра|лампа|патрон'), ('box', r'коробк'), ('frame', r'рамк'),
    ('panel', r'щит'), ('kup', r'куп|суп|уравнива'),
]


def type_id(code):
    return 't' + code.replace('.', '_').replace('-', '_')


def group_of_count(name):
    n = str(name).lower()
    for key, pat in GROUPS:
        if re.search(pat, n):
            return key
    return None


def read_types(path):
    """Листы рабочего .xlsx: шифр ЭОМ → спецификация на одну квартиру."""
    import openpyxl
    wb = openpyxl.load_workbook(path, data_only=True)
    out = {}
    for name in wb.sheetnames:
        if name in ('Сводная', 'Лист1'):
            continue
        ws = wb[name]
        hdr = ' '.join(str(ws['B6'].value or '').split())
        m = re.search(r'№?\s*(ФП-[\d.С]+?)-ЭОМ', hdr)
        if not m:
            continue
        items, sect = [], ''
        for row in ws.iter_rows(min_row=11, max_row=ws.max_row, max_col=7):
            pos, nm, unit, qty = row[0].value, row[1].value, row[5].value, row[6].value
            if nm and not pos:
                s = str(nm).strip()
                if 'Итого' not in s:
                    sect = re.sub(r'^\d+\.\s*', '', s)
                continue
            if nm and pos and isinstance(qty, (int, float)) and qty:
                items.append({
                    'p': str(pos).strip(),
                    'n': ' '.join(str(nm).split()),
                    'u': str(unit or 'шт.').strip().rstrip('.'),
                    'q': qty, 's': sect,
                })
        if items:
            out[m.group(1)] = {'hdr': hdr, 'items': items}
    return out


def read_walk(path):
    """Резервная копия обхода → итоги по квартирам. Фотографии игнорируем."""
    with open(path, encoding='utf-8') as f:
        src = json.load(f)
    cfg, data = src.get('cfg') or {}, src.get('data') or {}
    positions = cfg.get('positions') or []
    # раздел позиции («Санузел») — это зона; в санузлах работа обычно уходит
    # вперёд, и без разделения готовые санузлы не видны за незакрытой квартирой
    def zone_of(p):
        return 'wc' if re.search(r'санузел|с/у|ванн', str(p.get('g') or ''), re.I) else 'flat' 
    cmap = {c['id']: group_of_count(c['n']) for c in (cfg.get('count') or [])}
    crews = [{'id': w['id'], 'n': w['n']} for w in (cfg.get('crews') or [])]

    flats, total = {}, 0
    for bid, d in data.items():
        out = {}
        for num, r in d.items():
            if not r:
                continue
            rec = {}
            st = r.get('st') or {}
            vals = [st.get(p['id']) for p in positions]
            filled = [v for v in vals if v is not None]
            miss = [p['n'] for p in positions if st.get(p['id']) == 0]
            if filled:
                rec['c'] = 2 if miss else 1
            if miss:
                rec['miss'] = ', '.join(miss)
            # полный расклад приёмки: он же даёт готовность по зонам
            acc = {pid: v for pid, v in st.items() if v is not None}
            if acc:
                rec['st'] = acc
            # раньше в обходе было два поля; теперь одно, старое дописываем
            left = (r.get('left') or '').strip()
            note = (r.get('note') or '').strip()
            if note:
                left = (left + '\n' + note) if left else note
            if left:
                rec['left'] = left
            if r.get('crit'):
                rec['crit'] = True
            if r.get('fix'):
                rec['fix'] = r['fix']

            q = {}
            for cid, v in (r.get('q') or {}).items():
                k = cmap.get(cid)
                if k and v:
                    q[k] = q.get(k, 0) + v
            if q:
                rec['q'] = q

            # по бригадам — из журнала нажатий
            w = {}
            for e in (r.get('lg') or []):
                k = cmap.get(e.get('c'))
                if not k:
                    continue
                wid = e.get('w') or '?'
                w.setdefault(wid, {})
                w[wid][k] = w[wid].get(k, 0) + e.get('d', 0)
            for wid in list(w):
                w[wid] = {k: v for k, v in w[wid].items() if v > 0}
                if not w[wid]:
                    del w[wid]
            if not w and q:
                w['?'] = dict(q)          # старые записи без журнала
            if w:
                rec['w'] = w

            if rec:
                out[num] = rec
                total += 1
        if out:
            flats[bid] = out
    return flats, crews, total


def accepted_positions(path):
    """Позиции приёмки с зонами — по ним считается готовность квартиры и санузла."""
    with open(path, encoding='utf-8') as f:
        cfg = (json.load(f).get('cfg') or {})
    out = []
    for p in (cfg.get('positions') or []):
        z = 'wc' if re.search(r'санузел|с/у|ванн', str(p.get('g') or ''), re.I) else 'flat'
        out.append({'id': p['id'], 'n': p['n'], 'z': z})
    return out


def main():
    buildings, preset, missing = [], {}, []
    for b in BUILDINGS:
        path = os.path.join(PROJECTS_DIR, b['xlsx'])
        if not os.path.exists(path):
            sys.exit('Не найден файл проекта: ' + path)
        sheets = read_types(path)
        # 1.С и 1.1 могут лежать в файле соседнего корпуса
        for other in BUILDINGS:
            if other is b:
                continue
            op = os.path.join(PROJECTS_DIR, other['xlsx'])
            if os.path.exists(op):
                for code, v in read_types(op).items():
                    sheets.setdefault(code, v)

        types = []
        for code in b['types']:
            if code not in sheets:
                missing.append(b['name'] + ': ' + code)
                continue
            types.append({
                'id': type_id(code), 'n': code,
                'hdr': sheets[code]['hdr'], 'flats': [],
                'items': sheets[code]['items'],
            })
        buildings.append({'id': b['id'], 'name': b['name'], 'types': types})

        p = json.loads(json.dumps(b['preset']))
        p['stack'] = [type_id(c) if c else '' for c in p['stack']]
        for ex in p.get('ex', []):
            ex['stack'] = [type_id(c) if c else '' for c in ex['stack']]
        preset[b['id']] = p

    # приход раскладываем на настоящие ключи ведомости «название|единица»
    recv = {}
    for bid, items in RECEIVED.items():
        b = [x for x in buildings if x['id'] == bid]
        if not b:
            continue
        keys = {}
        for t in b[0]['types']:
            for it in t['items']:
                keys.setdefault(it['n'] + '|' + it['u'], it['n'])
        out, unmatched = {}, []
        for prefix, qty in items.items():
            hit = [k for k in keys if keys[k].startswith(prefix)]
            if len(hit) == 1:
                out[hit[0]] = qty
            else:
                unmatched.append((prefix, len(hit)))
        if out:
            recv[bid] = out
        if unmatched:
            print('приход, не сопоставлено (%s):' % bid, unmatched)
        print('приход (%s): строк %d' % (bid, len(out)))

    pkg = {
        't': 'uch-object', 'v': 1,
        'recv': recv,
        'object': OBJECT_NAME,
        'buildings': buildings,
        'preset': preset,
    }

    if WALK_BACKUP and os.path.exists(WALK_BACKUP):
        flats, crews, total = read_walk(WALK_BACKUP)
        pkg['walk'] = flats
        pkg['crews'] = crews
        pkg['walkName'] = os.path.basename(WALK_BACKUP)
        pkg['acc'] = accepted_positions(WALK_BACKUP)
        print('данные обхода: квартир %d, бригад %d' % (total, len(crews)))
    else:
        print('данные обхода не подключены')

    with open(OUT, 'w', encoding='utf-8') as f:
        json.dump(pkg, f, ensure_ascii=False, separators=(',', ':'))

    print('готово:', OUT, '(%.1f МБ)' % (os.path.getsize(OUT) / 1048576))
    for b in buildings:
        c = preset[b['id']]
        n = (c['to'] - c['from'] + 1) * c['per']
        for ex in c.get('ex', []):
            n += (ex['to'] - ex['from'] + 1) * (ex['per'] - c['per'])
        print('  %s: типов %d, квартир %d' % (b['name'], len(b['types']), n))
    if missing:
        print('НЕ НАЙДЕНЫ ЛИСТЫ:', ', '.join(missing))


if __name__ == '__main__':
    main()
